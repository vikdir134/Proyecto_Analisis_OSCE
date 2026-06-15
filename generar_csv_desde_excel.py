from pathlib import Path
from openpyxl import load_workbook
import csv
import re

# ==============================
# RUTAS
# ==============================

BASE_DIR = Path(__file__).resolve().parent
CARPETA_CSV = BASE_DIR / "CSV"

ENTRADA_XLSX = CARPETA_CSV / "CONSOLIDADO_CONTRATOS_2022_2025.xlsx"

SALIDA_CSV_UNA_COLUMNA = CARPETA_CSV / "CONSOLIDADO_CONTRATOS_2022_2025_UNA_COLUMNA.csv"
REPORTE = CARPETA_CSV / "REPORTE_CSV_UNA_COLUMNA.txt"


# ==============================
# FUNCIONES
# ==============================

def limpiar_valor(valor):
    """
    Limpia solo caracteres que rompen filas:
    - saltos de línea
    - tabulaciones
    - espacios múltiples

    NO elimina comas.
    NO elimina punto y coma.
    """
    if valor is None:
        return ""

    valor = str(valor)

    valor = valor.replace("\n", " ")
    valor = valor.replace("\r", " ")
    valor = valor.replace("\t", " ")

    valor = re.sub(r"\s+", " ", valor)

    return valor.strip()


def fila_vacia(fila):
    return all(celda is None or str(celda).strip() == "" for celda in fila)


def generar_csv_una_columna():
    if not ENTRADA_XLSX.exists():
        print(f"No existe el archivo de entrada: {ENTRADA_XLSX}")
        return

    print(f"Leyendo archivo Excel final:")
    print(ENTRADA_XLSX)

    wb = load_workbook(
        ENTRADA_XLSX,
        read_only=True,
        data_only=True
    )

    ws = wb.active

    total_filas_leidas = 0
    total_filas_escritas = 0
    total_filas_vacias = 0
    filas_con_pipe_interno = 0

    with open(
        SALIDA_CSV_UNA_COLUMNA,
        "w",
        newline="",
        encoding="utf-8-sig"
    ) as archivo_salida:

        # Usamos delimiter=";" solo para que Excel español no parta mal el archivo.
        # Pero escribimos UNA SOLA COLUMNA por fila.
        writer = csv.writer(
            archivo_salida,
            delimiter=";",
            quotechar='"',
            quoting=csv.QUOTE_ALL,
            lineterminator="\n"
        )

        for fila in ws.iter_rows(values_only=True):
            total_filas_leidas += 1

            if fila_vacia(fila):
                total_filas_vacias += 1
                continue

            valores_limpios = []

            for celda in fila:
                valor = limpiar_valor(celda)

                # Validación: si algún dato tiene | interno, puede confundirse luego.
                # No lo cambiamos, solo lo reportamos.
                if "|" in valor:
                    filas_con_pipe_interno += 1

                valores_limpios.append(valor)

            # Aquí se concatenan las columnas con |
            linea_con_pipe = "|".join(valores_limpios)

            # Aquí se escribe como UNA SOLA COLUMNA en el CSV
            writer.writerow([linea_con_pipe])

            total_filas_escritas += 1

    wb.close()

    with open(REPORTE, "w", encoding="utf-8") as f:
        f.write("REPORTE CSV UNA COLUMNA\n")
        f.write("=" * 80)
        f.write("\n")
        f.write(f"Archivo entrada: {ENTRADA_XLSX}\n")
        f.write(f"Archivo salida: {SALIDA_CSV_UNA_COLUMNA}\n")
        f.write(f"Filas leídas: {total_filas_leidas}\n")
        f.write(f"Filas escritas: {total_filas_escritas}\n")
        f.write(f"Filas vacías omitidas: {total_filas_vacias}\n")
        f.write(f"Filas con símbolo | dentro de algún dato: {filas_con_pipe_interno}\n")

    print("\n==============================")
    print("CSV GENERADO")
    print("==============================")
    print(f"Archivo generado: {SALIDA_CSV_UNA_COLUMNA}")
    print(f"Reporte generado: {REPORTE}")
    print(f"Filas leídas: {total_filas_leidas}")
    print(f"Filas escritas: {total_filas_escritas}")
    print(f"Filas vacías omitidas: {total_filas_vacias}")
    print(f"Filas con símbolo | interno: {filas_con_pipe_interno}")

    print("\nListo. Al abrir este CSV en Excel, cada fila debería quedar en la columna A.")
    print("Dentro de la columna A, los campos estarán separados por |.")


if __name__ == "__main__":
    generar_csv_una_columna()