from pathlib import Path
from openpyxl import load_workbook
import csv
import json


BASE_DIR = Path(__file__).resolve().parent

CARPETA_2026 = (
    BASE_DIR
    / "data"
    / "raw"
    / "oece"
    / "2026"
)

REPORTE = CARPETA_2026 / "reporte_estructura_2026.txt"


def inspeccionar_xlsx(archivo):
    resultados = []

    try:
        wb = load_workbook(
            archivo,
            read_only=True,
            data_only=True
        )

        resultados.append(f"Tipo: Excel XLSX")
        resultados.append(f"Hojas: {wb.sheetnames}")

        for nombre_hoja in wb.sheetnames:
            ws = wb[nombre_hoja]

            resultados.append(f"  Hoja: {nombre_hoja}")
            resultados.append(f"  Filas aproximadas: {ws.max_row}")
            resultados.append(f"  Columnas aproximadas: {ws.max_column}")

            primera_fila = next(
                ws.iter_rows(
                    min_row=1,
                    max_row=1,
                    values_only=True
                ),
                []
            )

            encabezados = [
                str(valor).strip() if valor is not None else ""
                for valor in primera_fila
            ]

            resultados.append("  Encabezados:")

            for posicion, encabezado in enumerate(encabezados, start=1):
                resultados.append(
                    f"    {posicion}. {encabezado}"
                )

        wb.close()

    except Exception as error:
        resultados.append(f"ERROR leyendo Excel: {error}")

    return resultados


def detectar_delimitador(archivo):
    try:
        with open(
            archivo,
            "r",
            encoding="utf-8-sig",
            errors="replace"
        ) as f:
            muestra = f.read(10000)

        dialecto = csv.Sniffer().sniff(
            muestra,
            delimiters=",;|\t"
        )

        return dialecto.delimiter

    except Exception:
        return None


def inspeccionar_csv(archivo):
    resultados = []

    delimitador = detectar_delimitador(archivo)

    resultados.append("Tipo: CSV")
    resultados.append(
        f"Delimitador detectado: {repr(delimitador)}"
    )

    if delimitador is None:
        resultados.append(
            "No se pudo detectar el delimitador."
        )
        return resultados

    try:
        with open(
            archivo,
            "r",
            encoding="utf-8-sig",
            errors="replace",
            newline=""
        ) as f:
            reader = csv.reader(
                f,
                delimiter=delimitador
            )

            encabezados = next(reader, [])

            resultados.append(
                f"Cantidad de columnas: {len(encabezados)}"
            )

            resultados.append("Encabezados:")

            for posicion, encabezado in enumerate(encabezados, start=1):
                resultados.append(
                    f"  {posicion}. {encabezado}"
                )

            filas_contadas = 1

            for _ in reader:
                filas_contadas += 1

            resultados.append(
                f"Filas incluyendo encabezado: {filas_contadas}"
            )

    except Exception as error:
        resultados.append(
            f"ERROR leyendo CSV: {error}"
        )

    return resultados


def inspeccionar_json(archivo):
    resultados = []

    resultados.append("Tipo: JSON")

    try:
        with open(
            archivo,
            "r",
            encoding="utf-8-sig",
            errors="replace"
        ) as f:
            contenido = json.load(f)

        resultados.append(
            f"Estructura principal: {type(contenido).__name__}"
        )

        if isinstance(contenido, dict):
            resultados.append("Claves principales:")

            for clave in contenido.keys():
                resultados.append(f"  - {clave}")

        elif isinstance(contenido, list):
            resultados.append(
                f"Cantidad de elementos: {len(contenido)}"
            )

            if contenido and isinstance(contenido[0], dict):
                resultados.append(
                    "Campos del primer elemento:"
                )

                for clave in contenido[0].keys():
                    resultados.append(f"  - {clave}")

    except Exception as error:
        resultados.append(
            f"ERROR leyendo JSON: {error}"
        )

    return resultados


def parece_archivo_contratos(nombre):
    nombre = nombre.lower()

    palabras = [
        "contrato",
        "contracts",
        "award",
        "adjudic",
        "ocds"
    ]

    return any(
        palabra in nombre
        for palabra in palabras
    )


def main():
    if not CARPETA_2026.exists():
        print(
            f"No existe la carpeta: {CARPETA_2026}"
        )
        return

    archivos = sorted([
        archivo
        for archivo in CARPETA_2026.rglob("*")
        if archivo.is_file()
        and archivo.suffix.lower() in {
            ".xlsx",
            ".xls",
            ".csv",
            ".json"
        }
        and archivo.name != "estado_descargas_2026.json"
    ])

    if not archivos:
        print(
            "No se encontraron archivos extraídos "
            "XLSX, CSV o JSON."
        )
        return

    lineas = []

    lineas.append("REPORTE DE ARCHIVOS OECE 2026")
    lineas.append("=" * 100)
    lineas.append(
        f"Cantidad de archivos encontrados: {len(archivos)}"
    )
    lineas.append("")

    candidatos = []

    for archivo in archivos:
        ruta_relativa = archivo.relative_to(CARPETA_2026)

        lineas.append("=" * 100)
        lineas.append(f"ARCHIVO: {ruta_relativa}")
        lineas.append(
            f"Tamaño: {archivo.stat().st_size / 1024 / 1024:.2f} MB"
        )

        if parece_archivo_contratos(archivo.name):
            lineas.append(
                "POSIBLE ARCHIVO DE CONTRATOS: SÍ"
            )
            candidatos.append(str(ruta_relativa))
        else:
            lineas.append(
                "POSIBLE ARCHIVO DE CONTRATOS: NO DETERMINADO"
            )

        extension = archivo.suffix.lower()

        if extension == ".xlsx":
            detalles = inspeccionar_xlsx(archivo)

        elif extension == ".csv":
            detalles = inspeccionar_csv(archivo)

        elif extension == ".json":
            detalles = inspeccionar_json(archivo)

        else:
            detalles = [
                f"Formato no inspeccionado: {extension}"
            ]

        lineas.extend(detalles)
        lineas.append("")

    lineas.append("=" * 100)
    lineas.append("POSIBLES ARCHIVOS RELACIONADOS CON CONTRATOS")

    if candidatos:
        for candidato in candidatos:
            lineas.append(f"- {candidato}")
    else:
        lineas.append(
            "No se detectaron por nombre. "
            "Habrá que revisar los encabezados."
        )

    with open(
        REPORTE,
        "w",
        encoding="utf-8"
    ) as f:
        f.write("\n".join(lineas))

    print("Inspección terminada.")
    print(f"Archivos encontrados: {len(archivos)}")
    print(f"Reporte generado: {REPORTE}")

    print("\nPosibles archivos de contratos:")

    if candidatos:
        for candidato in candidatos:
            print(f"- {candidato}")
    else:
        print(
            "No se identificaron por nombre. "
            "Revisa los encabezados del reporte."
        )


if __name__ == "__main__":
    main()