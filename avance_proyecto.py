from pathlib import Path
from openpyxl import load_workbook, Workbook
import re

# ==============================
# RUTAS
# ==============================

BASE_DIR = Path(__file__).resolve().parent
CARPETA_CSV = BASE_DIR / "CSV"

SALIDA_XLSX = CARPETA_CSV / "CONSOLIDADO_CONTRATOS_2022_2025.xlsx"
REPORTE = CARPETA_CSV / "REPORTE_CONSOLIDADO.txt"


# ==============================
# FUNCIONES
# ==============================

def normalizar_columna(valor):
    if valor is None:
        return ""

    valor = str(valor).strip().upper()
    valor = re.sub(r"\s+", "_", valor)
    return valor


def fila_vacia(fila):
    return all(celda is None or str(celda).strip() == "" for celda in fila)


def obtener_columnas_reales(ws):
    """
    Lee la primera fila y detecta hasta dónde llegan las columnas reales.
    Evita columnas vacías extra por formato de Excel.
    """
    encabezado = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))

    columnas = list(encabezado)

    while columnas and (columnas[-1] is None or str(columnas[-1]).strip() == ""):
        columnas.pop()

    return columnas


def validar_columnas(archivos):
    columnas_base = None
    columnas_base_norm = None
    reporte = []

    for archivo in archivos:
        wb = load_workbook(archivo, read_only=True, data_only=True)
        ws = wb.active

        columnas_actuales = obtener_columnas_reales(ws)
        columnas_actuales_norm = [normalizar_columna(c) for c in columnas_actuales]

        wb.close()

        if columnas_base is None:
            columnas_base = columnas_actuales
            columnas_base_norm = columnas_actuales_norm

            reporte.append(f"ARCHIVO BASE: {archivo.name}")
            reporte.append("COLUMNAS BASE:")
            for i, col in enumerate(columnas_base, start=1):
                reporte.append(f"{i}. {col}")
            reporte.append("")
        else:
            if columnas_actuales_norm != columnas_base_norm:
                reporte.append("=" * 80)
                reporte.append(f"ERROR DE COLUMNAS EN: {archivo.name}")
                reporte.append("")

                max_len = max(len(columnas_base_norm), len(columnas_actuales_norm))

                for i in range(max_len):
                    base = columnas_base_norm[i] if i < len(columnas_base_norm) else "NO EXISTE"
                    actual = columnas_actuales_norm[i] if i < len(columnas_actuales_norm) else "NO EXISTE"

                    if base != actual:
                        reporte.append(f"Posición {i + 1}")
                        reporte.append(f"Base:   {base}")
                        reporte.append(f"Actual: {actual}")
                        reporte.append("")

                with open(REPORTE, "w", encoding="utf-8") as f:
                    f.write("\n".join(reporte))

                raise Exception(
                    f"Las columnas de {archivo.name} no coinciden. "
                    f"Revisa REPORTE_CONSOLIDADO.txt"
                )

    with open(REPORTE, "w", encoding="utf-8") as f:
        f.write("\n".join(reporte))

    return columnas_base


def consolidar_excel():
    archivos = sorted(CARPETA_CSV.glob("CONOSCE_CONTRATOS*.xlsx"))

    if not archivos:
        print("No se encontraron archivos CONOSCE_CONTRATOS*.xlsx en la carpeta CSV.")
        return

    print("Archivos encontrados:")
    for archivo in archivos:
        print(f"- {archivo.name}")

    print("\nValidando columnas...")
    columnas = validar_columnas(archivos)
    cantidad_columnas = len(columnas)

    print("Columnas iguales. Se empezará a anexar como tabla Excel.")

    # Crear nuevo Excel
    wb_salida = Workbook(write_only=True)
    ws_salida = wb_salida.create_sheet("CONSOLIDADO")

    # Escribir encabezado solo una vez
    ws_salida.append(columnas)

    total_filas_escritas = 0
    total_filas_vacias = 0
    resumen = []

    for archivo in archivos:
        print(f"\nProcesando: {archivo.name}")

        wb = load_workbook(archivo, read_only=True, data_only=True)
        ws = wb.active

        filas_leidas = 0
        filas_escritas = 0
        filas_vacias = 0

        # Desde fila 2 porque fila 1 es encabezado
        for fila in ws.iter_rows(
            min_row=2,
            max_col=cantidad_columnas,
            values_only=True
        ):
            filas_leidas += 1

            if fila_vacia(fila):
                filas_vacias += 1
                continue

            # Aquí NO se separa nada.
            # Cada celda se copia como celda.
            ws_salida.append(list(fila))

            filas_escritas += 1
            total_filas_escritas += 1

        wb.close()

        total_filas_vacias += filas_vacias

        print(f"Filas leídas: {filas_leidas}")
        print(f"Filas anexadas: {filas_escritas}")
        print(f"Filas vacías eliminadas: {filas_vacias}")

        resumen.append(
            f"{archivo.name} | Leídas: {filas_leidas} | "
            f"Anexadas: {filas_escritas} | Vacías eliminadas: {filas_vacias}"
        )

    print("\nGuardando consolidado XLSX...")
    wb_salida.save(SALIDA_XLSX)

    with open(REPORTE, "a", encoding="utf-8") as f:
        f.write("\n\nRESUMEN DE CONSOLIDACIÓN\n")
        f.write("=" * 80)
        f.write("\n")
        f.write("\n".join(resumen))
        f.write("\n")
        f.write(f"\nTOTAL FILAS ANEXADAS: {total_filas_escritas}")
        f.write(f"\nTOTAL FILAS VACÍAS ELIMINADAS: {total_filas_vacias}")
        f.write(f"\nTOTAL FILAS EN EXCEL FINAL INCLUYENDO ENCABEZADO: {total_filas_escritas + 1}")

    print("\n==============================")
    print("CONSOLIDADO GENERADO")
    print("==============================")
    print(f"Archivo generado: {SALIDA_XLSX}")
    print(f"Reporte generado: {REPORTE}")
    print(f"Total filas anexadas: {total_filas_escritas}")
    print(f"Total filas en Excel final incluyendo encabezado: {total_filas_escritas + 1}")


if __name__ == "__main__":
    consolidar_excel()