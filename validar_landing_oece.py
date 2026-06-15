from __future__ import annotations

import argparse
import re
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


# ==========================================================
# CONFIGURACIÓN
# ==========================================================

BASE_DIR = Path(__file__).resolve().parent

CARPETA_RAW = (
    BASE_DIR
    / "data"
    / "raw"
    / "oece"
)

CARPETA_CALIDAD = (
    BASE_DIR
    / "data"
    / "quality"
    / "oece"
)

CARPETA_CALIDAD.mkdir(
    parents=True,
    exist_ok=True
)

SALIDA_EXCEL = (
    CARPETA_CALIDAD
    / "VALIDACION_LANDING_OECE_2022_2026.xlsx"
)

SALIDA_TEXTO = (
    CARPETA_CALIDAD
    / "VALIDACION_LANDING_OECE_2022_2026.txt"
)

ANIOS = [
    2022,
    2023,
    2024,
    2025,
    2026
]


# ==========================================================
# HOJAS Y COLUMNAS CRÍTICAS
# ==========================================================

HOJAS_OBLIGATORIAS = {
    "Registros": [
        "Open Contracting ID",
        "Entrega compilada:ID de Entrega",
        "Entrega compilada:Licitación:ID de licitación",
    ],

    "Ent_Contratos": [
        "Open Contracting ID",
        "Entrega compilada:ID de Entrega",
        "Entrega compilada:Contratos:ID del Contrato",
        "Entrega compilada:Contratos:ID de Adjudicación",
        "Entrega compilada:Contratos:Valor:Monto",
        "Entrega compilada:Contratos:Valor:Moneda",
    ],

    "Ent_Adj_Proveedores": [
        "Open Contracting ID",
        "Entrega compilada:ID de Entrega",
        "Entrega compilada:Adjudicaciones:ID de Adjudicación",
        "Entrega compilada:Adjudicaciones:Proveedores:ID de Organización",
        "Entrega compilada:Adjudicaciones:Proveedores:Nombre de la Organización",
    ],
}


HOJAS_COMPLEMENTARIAS = {
    "Ent_Con_ArticulosContratados": [
        "Open Contracting ID",
        "Entrega compilada:Contratos:ID del Contrato",
        "Entrega compilada:Contratos:Artículos Contratados:ID",
    ],

    "Ent_Con_Documentos": [
        "Open Contracting ID",
        "Entrega compilada:Contratos:ID del Contrato",
        "Entrega compilada:Contratos:Documentos:URL",
    ],

    "Ent_PartesInvolucradas": [
        "Open Contracting ID",
        "Entrega compilada:Partes involucradas:ID de Entidad",
        "Entrega compilada:Partes involucradas:Identificador principal:ID",
    ],
}


# ==========================================================
# FUNCIONES GENERALES
# ==========================================================

def normalizar_texto(valor: Any) -> str:
    """
    Normaliza únicamente para comparar encabezados.
    No modifica los archivos originales.
    """

    if valor is None:
        return ""

    texto = str(valor)

    texto = texto.replace("\n", " ")
    texto = texto.replace("\r", " ")
    texto = texto.replace("\t", " ")

    texto = re.sub(
        r"\s+",
        " ",
        texto
    )

    return texto.strip()


def normalizar_comparacion(valor: Any) -> str:
    """
    Versión más flexible para comparar nombres de columnas.
    """

    return normalizar_texto(valor).casefold()


def convertir_mes(valor: Any) -> str:
    try:
        return str(int(valor)).zfill(2)
    except (TypeError, ValueError):
        return ""


def meses_esperados(anio: int) -> list[str]:
    """
    Para años cerrados espera 12 meses.

    Para el año actual espera hasta el mes actual.
    """

    hoy = datetime.now()

    if anio < hoy.year:
        ultimo_mes = 12

    elif anio == hoy.year:
        ultimo_mes = hoy.month

    else:
        ultimo_mes = 0

    return [
        str(mes).zfill(2)
        for mes in range(1, ultimo_mes + 1)
    ]


def obtener_archivos_mes(
    anio: int,
    mes: str
) -> list[Path]:

    carpeta_extraida = (
        CARPETA_RAW
        / str(anio)
        / mes
        / "extraido"
    )

    if not carpeta_extraida.exists():
        return []

    return sorted(
        archivo
        for archivo in carpeta_extraida.rglob("*.xlsx")
        if archivo.is_file()
        and not archivo.name.startswith("~$")
    )


def seleccionar_archivo_principal(
    archivos: list[Path]
) -> Path | None:
    """
    Cuando existen varios XLSX, selecciona el de mayor tamaño.
    """

    if not archivos:
        return None

    return max(
        archivos,
        key=lambda archivo: archivo.stat().st_size
    )


def validar_archivo_xlsx(archivo: Path) -> tuple[bool, str]:
    """
    Un XLSX válido internamente es un archivo ZIP.
    """

    if not archivo.exists():
        return False, "ARCHIVO_NO_EXISTE"

    if archivo.stat().st_size == 0:
        return False, "ARCHIVO_VACIO"

    if not zipfile.is_zipfile(archivo):
        return False, "NO_ES_XLSX_VALIDO"

    return True, "OK"


def leer_encabezados(ws) -> list[str]:
    primera_fila = next(
        ws.iter_rows(
            min_row=1,
            max_row=1,
            values_only=True
        ),
        ()
    )

    encabezados = [
        normalizar_texto(valor)
        for valor in primera_fila
    ]

    while encabezados and encabezados[-1] == "":
        encabezados.pop()

    return encabezados


def contar_filas_datos(ws) -> int:
    """
    Cuenta filas no vacías, excluyendo el encabezado.
    Funciona aunque ws.max_row aparezca como None.
    """

    total = 0

    for fila in ws.iter_rows(
        min_row=2,
        values_only=True
    ):
        if any(
            valor is not None
            and str(valor).strip() != ""
            for valor in fila
        ):
            total += 1

    return total


def buscar_columnas_faltantes(
    encabezados: list[str],
    columnas_esperadas: list[str]
) -> list[str]:

    encabezados_normalizados = {
        normalizar_comparacion(columna)
        for columna in encabezados
    }

    return [
        columna
        for columna in columnas_esperadas
        if normalizar_comparacion(columna)
        not in encabezados_normalizados
    ]


# ==========================================================
# VALIDACIÓN DE UN ARCHIVO
# ==========================================================

def validar_mes(
    anio: int,
    mes: str,
    contar_filas: bool
) -> tuple[
    dict[str, Any],
    list[dict[str, Any]],
    list[dict[str, Any]]
]:

    archivos = obtener_archivos_mes(
        anio,
        mes
    )

    resumen_mes = {
        "ANIO": anio,
        "MES": mes,
        "CANTIDAD_XLSX": len(archivos),
        "ARCHIVO_PRINCIPAL": "",
        "TAMANO_MB": 0,
        "HOJAS_ENCONTRADAS": 0,
        "HOJAS_OBLIGATORIAS_FALTANTES": 0,
        "COLUMNAS_CRITICAS_FALTANTES": 0,
        "FILAS_REGISTROS": None,
        "FILAS_CONTRATOS": None,
        "FILAS_PROVEEDORES": None,
        "ESTADO": "",
        "OBSERVACION": "",
    }

    detalle_hojas = []
    errores_columnas = []

    if not archivos:
        resumen_mes["ESTADO"] = "FALTANTE"
        resumen_mes["OBSERVACION"] = (
            "No existe un archivo XLSX extraído para este mes."
        )

        return (
            resumen_mes,
            detalle_hojas,
            errores_columnas
        )

    archivo = seleccionar_archivo_principal(
        archivos
    )

    if archivo is None:
        resumen_mes["ESTADO"] = "FALTANTE"
        resumen_mes["OBSERVACION"] = (
            "No se pudo seleccionar un archivo principal."
        )

        return (
            resumen_mes,
            detalle_hojas,
            errores_columnas
        )

    resumen_mes["ARCHIVO_PRINCIPAL"] = str(
        archivo.relative_to(CARPETA_RAW)
    )

    resumen_mes["TAMANO_MB"] = round(
        archivo.stat().st_size
        / 1024
        / 1024,
        2
    )

    valido, estado_archivo = validar_archivo_xlsx(
        archivo
    )

    if not valido:
        resumen_mes["ESTADO"] = "ERROR"
        resumen_mes["OBSERVACION"] = estado_archivo

        return (
            resumen_mes,
            detalle_hojas,
            errores_columnas
        )

    try:
        wb = load_workbook(
            archivo,
            read_only=True,
            data_only=True
        )

    except Exception as error:
        resumen_mes["ESTADO"] = "ERROR"
        resumen_mes["OBSERVACION"] = (
            f"No se pudo abrir: {error}"
        )

        return (
            resumen_mes,
            detalle_hojas,
            errores_columnas
        )

    resumen_mes["HOJAS_ENCONTRADAS"] = len(
        wb.sheetnames
    )

    hojas_a_validar = {
        **HOJAS_OBLIGATORIAS,
        **HOJAS_COMPLEMENTARIAS,
    }

    hojas_obligatorias_faltantes = [
        hoja
        for hoja in HOJAS_OBLIGATORIAS
        if hoja not in wb.sheetnames
    ]

    resumen_mes[
        "HOJAS_OBLIGATORIAS_FALTANTES"
    ] = len(hojas_obligatorias_faltantes)

    total_columnas_faltantes = 0

    for hoja, columnas_esperadas in hojas_a_validar.items():
        es_obligatoria = hoja in HOJAS_OBLIGATORIAS

        if hoja not in wb.sheetnames:
            detalle_hojas.append({
                "ANIO": anio,
                "MES": mes,
                "ARCHIVO": archivo.name,
                "HOJA": hoja,
                "TIPO_HOJA": (
                    "OBLIGATORIA"
                    if es_obligatoria
                    else "COMPLEMENTARIA"
                ),
                "EXISTE": "NO",
                "CANTIDAD_COLUMNAS": 0,
                "CANTIDAD_FILAS": None,
                "COLUMNAS_FALTANTES": len(
                    columnas_esperadas
                ),
                "ESTADO": (
                    "ERROR"
                    if es_obligatoria
                    else "ADVERTENCIA"
                ),
            })

            for columna in columnas_esperadas:
                errores_columnas.append({
                    "ANIO": anio,
                    "MES": mes,
                    "ARCHIVO": archivo.name,
                    "HOJA": hoja,
                    "COLUMNA_FALTANTE": columna,
                    "TIPO": (
                        "OBLIGATORIA"
                        if es_obligatoria
                        else "COMPLEMENTARIA"
                    ),
                })

            if es_obligatoria:
                total_columnas_faltantes += len(
                    columnas_esperadas
                )

            continue

        ws = wb[hoja]

        encabezados = leer_encabezados(ws)

        faltantes = buscar_columnas_faltantes(
            encabezados,
            columnas_esperadas
        )

        cantidad_filas = (
            contar_filas_datos(ws)
            if contar_filas
            else None
        )

        if hoja == "Registros":
            resumen_mes[
                "FILAS_REGISTROS"
            ] = cantidad_filas

        elif hoja == "Ent_Contratos":
            resumen_mes[
                "FILAS_CONTRATOS"
            ] = cantidad_filas

        elif hoja == "Ent_Adj_Proveedores":
            resumen_mes[
                "FILAS_PROVEEDORES"
            ] = cantidad_filas

        detalle_hojas.append({
            "ANIO": anio,
            "MES": mes,
            "ARCHIVO": archivo.name,
            "HOJA": hoja,
            "TIPO_HOJA": (
                "OBLIGATORIA"
                if es_obligatoria
                else "COMPLEMENTARIA"
            ),
            "EXISTE": "SI",
            "CANTIDAD_COLUMNAS": len(encabezados),
            "CANTIDAD_FILAS": cantidad_filas,
            "COLUMNAS_FALTANTES": len(faltantes),
            "ESTADO": (
                "OK"
                if not faltantes
                else (
                    "ERROR"
                    if es_obligatoria
                    else "ADVERTENCIA"
                )
            ),
        })

        for columna in faltantes:
            errores_columnas.append({
                "ANIO": anio,
                "MES": mes,
                "ARCHIVO": archivo.name,
                "HOJA": hoja,
                "COLUMNA_FALTANTE": columna,
                "TIPO": (
                    "OBLIGATORIA"
                    if es_obligatoria
                    else "COMPLEMENTARIA"
                ),
            })

        if es_obligatoria:
            total_columnas_faltantes += len(
                faltantes
            )

    wb.close()

    resumen_mes[
        "COLUMNAS_CRITICAS_FALTANTES"
    ] = total_columnas_faltantes

    observaciones = []

    if len(archivos) > 1:
        observaciones.append(
            f"Se encontraron {len(archivos)} archivos XLSX."
        )

    if hojas_obligatorias_faltantes:
        observaciones.append(
            "Faltan hojas obligatorias: "
            + ", ".join(hojas_obligatorias_faltantes)
        )

    if total_columnas_faltantes > 0:
        observaciones.append(
            "Existen columnas críticas faltantes."
        )

    if hojas_obligatorias_faltantes:
        resumen_mes["ESTADO"] = "ERROR"

    elif total_columnas_faltantes > 0:
        resumen_mes["ESTADO"] = "ERROR"

    elif len(archivos) > 1:
        resumen_mes["ESTADO"] = "ADVERTENCIA"

    else:
        resumen_mes["ESTADO"] = "OK"

    resumen_mes["OBSERVACION"] = " ".join(
        observaciones
    )

    return (
        resumen_mes,
        detalle_hojas,
        errores_columnas
    )


# ==========================================================
# PROCESO PRINCIPAL
# ==========================================================

def validar_landing(contar_filas: bool) -> None:
    if not CARPETA_RAW.exists():
        print(
            "No existe la carpeta:"
        )
        print(CARPETA_RAW)
        return

    resumen_meses = []
    detalle_hojas = []
    columnas_faltantes = []

    meses_no_esperados = []

    print("========================================")
    print("VALIDACIÓN DE LANDING ZONE OECE")
    print("========================================")

    for anio in ANIOS:
        esperados = meses_esperados(anio)

        print(f"\nAño {anio}")
        print(
            f"Meses esperados: {len(esperados)}"
        )

        for mes in esperados:
            print(
                f"Validando {anio}-{mes}..."
            )

            resumen, detalle, faltantes = validar_mes(
                anio,
                mes,
                contar_filas
            )

            resumen_meses.append(resumen)
            detalle_hojas.extend(detalle)
            columnas_faltantes.extend(faltantes)

        carpeta_anio = (
            CARPETA_RAW
            / str(anio)
        )

        if carpeta_anio.exists():
            carpetas_mes = [
                carpeta.name
                for carpeta in carpeta_anio.iterdir()
                if carpeta.is_dir()
                and carpeta.name.isdigit()
            ]

            for mes_encontrado in carpetas_mes:
                mes_normalizado = convertir_mes(
                    mes_encontrado
                )

                if (
                    mes_normalizado
                    and mes_normalizado not in esperados
                ):
                    meses_no_esperados.append({
                        "ANIO": anio,
                        "MES": mes_normalizado,
                        "RUTA": str(
                            (
                                carpeta_anio
                                / mes_encontrado
                            ).relative_to(
                                CARPETA_RAW
                            )
                        ),
                        "OBSERVACION": (
                            "Existe una carpeta mensual que "
                            "no estaba dentro del periodo esperado."
                        ),
                    })

    df_resumen = pd.DataFrame(
        resumen_meses
    )

    df_detalle = pd.DataFrame(
        detalle_hojas
    )

    df_columnas = pd.DataFrame(
        columnas_faltantes,
        columns=[
            "ANIO",
            "MES",
            "ARCHIVO",
            "HOJA",
            "COLUMNA_FALTANTE",
            "TIPO",
        ]
    )

    df_meses_extra = pd.DataFrame(
        meses_no_esperados,
        columns=[
            "ANIO",
            "MES",
            "RUTA",
            "OBSERVACION",
        ]
    )

    total_meses = len(df_resumen)

    meses_ok = int(
        df_resumen["ESTADO"]
        .eq("OK")
        .sum()
    )

    meses_advertencia = int(
        df_resumen["ESTADO"]
        .eq("ADVERTENCIA")
        .sum()
    )

    meses_error = int(
        df_resumen["ESTADO"]
        .eq("ERROR")
        .sum()
    )

    meses_faltantes = int(
        df_resumen["ESTADO"]
        .eq("FALTANTE")
        .sum()
    )

    total_contratos = (
        pd.to_numeric(
            df_resumen["FILAS_CONTRATOS"],
            errors="coerce"
        )
        .fillna(0)
        .sum()
    )

    resumen_general = pd.DataFrame([
        {
            "INDICADOR": "Meses esperados",
            "VALOR": total_meses,
        },
        {
            "INDICADOR": "Meses correctos",
            "VALOR": meses_ok,
        },
        {
            "INDICADOR": "Meses con advertencia",
            "VALOR": meses_advertencia,
        },
        {
            "INDICADOR": "Meses con error",
            "VALOR": meses_error,
        },
        {
            "INDICADOR": "Meses faltantes",
            "VALOR": meses_faltantes,
        },
        {
            "INDICADOR": "Columnas faltantes detectadas",
            "VALOR": len(df_columnas),
        },
        {
            "INDICADOR": "Filas de contratos detectadas",
            "VALOR": int(total_contratos),
        },
        {
            "INDICADOR": "Conteo detallado ejecutado",
            "VALOR": (
                "SI"
                if contar_filas
                else "NO"
            ),
        },
    ])

    print("\nGuardando reporte Excel...")

    with pd.ExcelWriter(
        SALIDA_EXCEL,
        engine="openpyxl"
    ) as writer:

        resumen_general.to_excel(
            writer,
            sheet_name="RESUMEN_GENERAL",
            index=False
        )

        df_resumen.to_excel(
            writer,
            sheet_name="RESUMEN_MESES",
            index=False
        )

        df_detalle.to_excel(
            writer,
            sheet_name="DETALLE_HOJAS",
            index=False
        )

        df_columnas.to_excel(
            writer,
            sheet_name="COLUMNAS_FALTANTES",
            index=False
        )

        df_meses_extra.to_excel(
            writer,
            sheet_name="MESES_NO_ESPERADOS",
            index=False
        )

    with open(
        SALIDA_TEXTO,
        "w",
        encoding="utf-8"
    ) as archivo:

        archivo.write(
            "VALIDACIÓN LANDING ZONE OECE 2022-2026\n"
        )

        archivo.write("=" * 80)
        archivo.write("\n")

        archivo.write(
            f"Fecha de ejecución: "
            f"{datetime.now().isoformat()}\n"
        )

        archivo.write(
            f"Meses esperados: {total_meses}\n"
        )

        archivo.write(
            f"Meses correctos: {meses_ok}\n"
        )

        archivo.write(
            f"Meses con advertencia: "
            f"{meses_advertencia}\n"
        )

        archivo.write(
            f"Meses con error: {meses_error}\n"
        )

        archivo.write(
            f"Meses faltantes: {meses_faltantes}\n"
        )

        archivo.write(
            f"Columnas faltantes: "
            f"{len(df_columnas)}\n"
        )

        archivo.write(
            f"Filas de contratos detectadas: "
            f"{int(total_contratos)}\n"
        )

        archivo.write("\nDETALLE DE PROBLEMAS\n")

        problemas = df_resumen[
            df_resumen["ESTADO"] != "OK"
        ]

        if problemas.empty:
            archivo.write(
                "No se detectaron problemas.\n"
            )

        else:
            for _, fila in problemas.iterrows():
                archivo.write(
                    f"- {fila['ANIO']}-{fila['MES']} | "
                    f"{fila['ESTADO']} | "
                    f"{fila['OBSERVACION']}\n"
                )

    print("\n========================================")
    print("RESULTADO")
    print("========================================")
    print(f"Meses esperados: {total_meses}")
    print(f"Correctos: {meses_ok}")
    print(
        f"Con advertencia: {meses_advertencia}"
    )
    print(f"Con error: {meses_error}")
    print(f"Faltantes: {meses_faltantes}")
    print(
        f"Columnas faltantes: "
        f"{len(df_columnas)}"
    )
    print(
        f"Filas de contratos detectadas: "
        f"{int(total_contratos)}"
    )
    print(f"\nExcel: {SALIDA_EXCEL}")
    print(f"Reporte: {SALIDA_TEXTO}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Valida la Landing Zone histórica "
            "de archivos OCDS/OECE."
        )
    )

    parser.add_argument(
        "--rapido",
        action="store_true",
        help=(
            "Valida archivos, hojas y columnas, "
            "pero no cuenta todas las filas."
        )
    )

    argumentos = parser.parse_args()

    validar_landing(
        contar_filas=not argumentos.rapido
    )


if __name__ == "__main__":
    main()